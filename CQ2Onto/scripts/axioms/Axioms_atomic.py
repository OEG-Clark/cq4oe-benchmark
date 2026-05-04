import sys
import os
import json
import tempfile
from collections import defaultdict
from owlready2 import *
import rdflib
from rdflib import RDF, RDFS, OWL, URIRef
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

_EXT_TO_FMT = {
    ".owl": "xml",
    ".rdf": "xml",
    ".xml": "xml",
    ".ttl": "turtle",
    ".turtle": "turtle",
    ".nt": "nt",
    ".n3": "n3",
    ".jsonld": "json-ld",
    ".json": "json-ld",
    ".trig": "trig",
}

_FALLBACK_FORMATS = ["xml", "turtle", "n3", "nt", "json-ld", "trig"]


def _detect_rdf_format(path):
    ext = os.path.splitext(path)[1].lower()
    return _EXT_TO_FMT.get(ext)


def _parse_rdf_with_fallback(path):
    primary = _detect_rdf_format(path)
    formats = [primary] if primary else []
    formats += [f for f in _FALLBACK_FORMATS if f != primary]

    last_err = None
    for fmt in formats:
        try:
            g = rdflib.Graph()
            g.parse(path, format=fmt)
            return g, fmt
        except Exception as e:
            last_err = e
    raise RuntimeError(f"Could not parse '{path}' in any known RDF format. "
                       f"Last error: {last_err}")


def _ensure_rdfxml_for_owlready(path):
    fmt = _detect_rdf_format(path)
    if fmt == "xml":
        return path, None
    g, used_fmt = _parse_rdf_with_fallback(path)
    tmp = tempfile.NamedTemporaryFile(
        suffix=".owl", delete=False, mode="wb")
    tmp.close()
    g.serialize(destination=tmp.name, format="xml")
    print(f"  Converted '{path}' (format={used_fmt}) → RDF/XML temp file "
          f"for owlready2: {tmp.name}")
    return tmp.name, tmp.name

_LABEL_LOOKUP: dict = {}


def _clean_name(name):
    name = str(name)
    if '#' in name:
        name = name.split('#')[-1]
    if '.' in name and not name.startswith('xsd'):
        name = name.split('.')[-1]
    return name


def _to_display_name(fragment: str) -> str:
    if not fragment:
        return fragment
    return _LABEL_LOOKUP.get(fragment, fragment)


def _get_prop_name(prop):
    if isinstance(prop, Inverse):
        inner = getattr(prop.property, 'name', str(prop.property))
        return _to_display_name(_clean_name(inner)) + "⁻"
    return _to_display_name(_clean_name(getattr(prop, 'name', str(prop))))


def _python_type_to_xsd(t):
    return {
        int: "xsd:integer",
        float: "xsd:float",
        str: "xsd:string",
        bool: "xsd:boolean",
    }.get(t, _clean_name(str(t)))


def _is_literal_like(v):
    return isinstance(v, (str, int, float, bool))


def _literal_to_json(v):
    if isinstance(v, (str, int, float, bool)):
        return v
    return str(v)


def _looks_like_data_expr(v):
    if v is None:
        return False
    if isinstance(v, ConstrainedDatatype):
        return True
    if isinstance(v, (ThingClass, And, Or, Not, OneOf, Restriction)):
        return False
    if isinstance(v, type):
        return True
    if _is_literal_like(v):
        return True
    return False


def _get_filler(expr):
    return getattr(expr, "value", None) or getattr(expr, "filler", None)

def to_dl(expr):
    if expr is None:
        return "⊤"
    if expr is Thing:
        return "⊤"
    if expr is Nothing:
        return "⊥"

    if isinstance(expr, ThingClass) and not isinstance(expr, (And, Or, Not, OneOf, Restriction)):
        return _to_display_name(_clean_name(expr.name))

    if isinstance(expr, And):
        return "(" + " ⊓ ".join(to_dl(c) for c in expr.Classes) + ")"

    if isinstance(expr, Or):
        return "(" + " ⊔ ".join(to_dl(c) for c in expr.Classes) + ")"

    if isinstance(expr, Not):
        return "¬" + to_dl(expr.Class)

    if isinstance(expr, OneOf):
        return "{" + ", ".join(_to_display_name(_clean_name(getattr(i, 'name', str(i)))) for i in expr.instances) + "}"

    if isinstance(expr, Restriction):
        p = _get_prop_name(expr.property)
        v = _get_filler(expr)
        if expr.type == SOME:
            return f"∃{p}.{to_dl(v)}"
        elif expr.type == ONLY:
            return f"∀{p}.{to_dl(v)}"
        elif expr.type == VALUE:
            if _is_literal_like(v):
                return f"∃{p}.{{\"{v}\"}}"
            vname = getattr(v, 'name', v)
            return f"∃{p}.{{{_to_display_name(_clean_name(vname))}}}"
        elif expr.type == MIN:
            return f"≥{expr.cardinality} {p}.{to_dl(v) if v else '⊤'}"
        elif expr.type == MAX:
            return f"≤{expr.cardinality} {p}.{to_dl(v) if v else '⊤'}"
        elif expr.type == EXACTLY:
            return f"={expr.cardinality} {p}.{to_dl(v) if v else '⊤'}"
        elif expr.type == HAS_SELF:
            return f"∃{p}.Self"

    if isinstance(expr, Inverse):
        return _get_prop_name(expr.property) + "⁻"

    if isinstance(expr, ConstrainedDatatype):
        base = _clean_name(str(expr.base_datatype))
        facets = ", ".join(f"{k}={v}" for k, v in expr.constrains)
        return f"{base}[{facets}]"

    if isinstance(expr, type):
        return _python_type_to_xsd(expr)

    if hasattr(expr, 'name'):
        return _to_display_name(_clean_name(expr.name))

    return _to_display_name(_clean_name(str(expr)))

def expr_to_struct(expr):
    if expr is None:
        return {"expr_type": "Top"}
    if expr is Thing:
        return {"expr_type": "Top"}
    if expr is Nothing:
        return {"expr_type": "Bottom"}

    if isinstance(expr, ThingClass) and not isinstance(expr, (And, Or, Not, OneOf, Restriction)):
        return {"expr_type": "Class", "name": _to_display_name(_clean_name(expr.name))}

    if isinstance(expr, And):
        return {
            "expr_type": "ObjectIntersectionOf",
            "operands": [expr_to_struct(c) for c in expr.Classes]
        }
    if isinstance(expr, Or):
        return {
            "expr_type": "ObjectUnionOf",
            "operands": [expr_to_struct(c) for c in expr.Classes]
        }
    if isinstance(expr, Not):
        return {
            "expr_type": "ObjectComplementOf",
            "operand": expr_to_struct(expr.Class)
        }
    if isinstance(expr, OneOf):
        return {
            "expr_type": "ObjectOneOf",
            "individuals": [_to_display_name(_clean_name(getattr(i, 'name', str(i)))) for i in expr.instances]
        }
    if isinstance(expr, Restriction):
        prop_name = _get_prop_name(expr.property)
        v = _get_filler(expr)
        is_data_side = _looks_like_data_expr(v)
        value_struct = expr_to_struct(v) if v is not None else None

        if expr.type == SOME:
            return {
                "expr_type": "DataSomeValuesFrom" if is_data_side else "ObjectSomeValuesFrom",
                "property": prop_name,
                "filler": value_struct,
            }
        elif expr.type == ONLY:
            return {
                "expr_type": "DataAllValuesFrom" if is_data_side else "ObjectAllValuesFrom",
                "property": prop_name,
                "filler": value_struct,
            }
        elif expr.type == VALUE:
            if is_data_side or _is_literal_like(v):
                return {
                    "expr_type": "DataHasValue",
                    "property": prop_name,
                    "value": _literal_to_json(v),
                }
            return {
                "expr_type": "ObjectHasValue",
                "property": prop_name,
                "value": _to_display_name(_clean_name(getattr(v, 'name', str(v)))),
            }
        elif expr.type == MIN:
            return {
                "expr_type": "DataMinCardinality" if is_data_side else "ObjectMinCardinality",
                "property": prop_name,
                "cardinality": expr.cardinality,
                "filler": value_struct if value_struct is not None else {"expr_type": "Top"},
            }
        elif expr.type == MAX:
            return {
                "expr_type": "DataMaxCardinality" if is_data_side else "ObjectMaxCardinality",
                "property": prop_name,
                "cardinality": expr.cardinality,
                "filler": value_struct if value_struct is not None else {"expr_type": "Top"},
            }
        elif expr.type == EXACTLY:
            return {
                "expr_type": "DataExactCardinality" if is_data_side else "ObjectExactCardinality",
                "property": prop_name,
                "cardinality": expr.cardinality,
                "filler": value_struct if value_struct is not None else {"expr_type": "Top"},
            }
        elif expr.type == HAS_SELF:
            return {"expr_type": "ObjectHasSelf", "property": prop_name}

    if isinstance(expr, ConstrainedDatatype):
        return {
            "expr_type": "DatatypeRestriction",
            "base": _clean_name(str(expr.base_datatype)),
            "facets": [{"facet": str(k), "value": v} for k, v in expr.constrains],
        }

    if isinstance(expr, type):
        return {"expr_type": "Datatype", "name": _python_type_to_xsd(expr)}

    if hasattr(expr, 'name'):
        return {"expr_type": "NamedEntity", "name": _to_display_name(_clean_name(expr.name))}

    return {"expr_type": "LiteralOrRaw", "value": str(expr)}

CHAR_URIS = {
    "Functional": OWL.FunctionalProperty,
    "InverseFunctional": OWL.InverseFunctionalProperty,
    "Transitive": OWL.TransitiveProperty,
    "Symmetric": OWL.SymmetricProperty,
    "Asymmetric": OWL.AsymmetricProperty,
    "Reflexive": OWL.ReflexiveProperty,
    "Irreflexive": OWL.IrreflexiveProperty,
}


def get_characteristics(prop_iri, graph):
    return [
        name for name, owl_type in CHAR_URIS.items()
        if (prop_iri, RDF.type, owl_type) in graph
    ]

def _flatten_and(expr):

    if isinstance(expr, And):
        out = []
        for c in expr.Classes:
            out.extend(_flatten_and(c))
        return out
    return [expr]


def _flatten_struct_intersection(struct):

    if struct and struct.get("expr_type") == "ObjectIntersectionOf":
        out = []
        for op in struct.get("operands", []):
            out.extend(_flatten_struct_intersection(op))
        return out
    return [struct]


def extract_tbox_axioms(ontology_path):

    owl_path, temp_to_cleanup = _ensure_rdfxml_for_owlready(ontology_path)
    try:
        onto = get_ontology(owl_path).load()
    finally:
        pass

    rdf_graph, fmt_used = _parse_rdf_with_fallback(ontology_path)
    print(f"  Parsed '{ontology_path}' as RDF format='{fmt_used}'")

    from rdflib import Literal as _RDFLiteral
    name_to_label = {}

    def _resolve_label(uri_node) -> str:
        labels = list(rdf_graph.objects(uri_node, RDFS.label))
        for lb in labels:
            if isinstance(lb, _RDFLiteral) and getattr(lb, "language", None) == "en":
                return str(lb).strip()
        for lb in labels:
            if isinstance(lb, _RDFLiteral) and getattr(lb, "language", None) in (None, ""):
                return str(lb).strip()
        for lb in labels:
            if isinstance(lb, _RDFLiteral):
                return str(lb).strip()
        s = str(uri_node)
        if "#" in s:
            return s.split("#")[-1]
        return s.rstrip("/").split("/")[-1]
    
    for s in rdf_graph.subjects(RDF.type, OWL.Class):
        if isinstance(s, URIRef):
            name = str(s).split("#")[-1] if "#" in str(s) else str(s).rstrip("/").split("/")[-1]
            name_to_label[name] = _resolve_label(s)
    for s in rdf_graph.subjects(RDF.type, OWL.ObjectProperty):
        if isinstance(s, URIRef):
            name = str(s).split("#")[-1] if "#" in str(s) else str(s).rstrip("/").split("/")[-1]
            name_to_label[name] = _resolve_label(s)
    for s in rdf_graph.subjects(RDF.type, OWL.DatatypeProperty):
        if isinstance(s, URIRef):
            name = str(s).split("#")[-1] if "#" in str(s) else str(s).rstrip("/").split("/")[-1]
            name_to_label[name] = _resolve_label(s)

    global _LABEL_LOOKUP
    _LABEL_LOOKUP = dict(name_to_label)

    terms = defaultdict(lambda: {"term_type": None, "axioms": []})
    stats = defaultdict(int)
    seen_disjoint_classes = set()
    seen_disjoint_props = set()
    seen_subclass = set()
    seen_subprop = set()

    def add_axiom(term_name, term_type, axiom_type, dl_str, subject, object_,
                  lhs_struct=None, rhs_struct=None):
        if terms[term_name]["term_type"] is None:
            terms[term_name]["term_type"] = term_type
        terms[term_name]["axioms"].append({
            "axiom_type": axiom_type,
            "dl": dl_str,
            "subject": subject,
            "object": object_,
            "lhs_struct": lhs_struct,
            "rhs_struct": rhs_struct,
        })
        stats["total_axioms"] += 1
        stats[axiom_type] += 1

    def _add_subclass_decomposed(class_name, sup_expr):
        operands = _flatten_and(sup_expr)
        for op in operands:
            if op is Thing:
                continue
            dl = to_dl(op)
            key = (class_name, "SubClassOf", dl)
            if key in seen_subclass:
                continue
            seen_subclass.add(key)
            add_axiom(
                term_name=class_name,
                term_type="Class",
                axiom_type="SubClassOf",
                dl_str=f"{class_name} ⊑ {dl}",
                subject=class_name,
                object_=dl,
                lhs_struct={"expr_type": "Class", "name": class_name},
                rhs_struct=expr_to_struct(op),
            )

    for cls in onto.classes():
        name = _to_display_name(_clean_name(cls.name))
        stats["total_classes"] += 1
        has_equiv = False

        add_axiom(
            term_name=name,
            term_type="Class",
            axiom_type="Declaration",
            dl_str=f"Declaration: {name} (Class)",
            subject=name,
            object_="Class",
            lhs_struct={"expr_type": "Class", "name": name},
            rhs_struct={"expr_type": "Declaration",
                        "term_type": "Class"},
        )

        for eq in cls.equivalent_to:
            eq_dl = to_dl(eq)
            eq_struct = expr_to_struct(eq)
            add_axiom(
                term_name=name,
                term_type="Class",
                axiom_type="EquivalentClasses",
                dl_str=f"{name} ≡ {eq_dl}",
                subject=name,
                object_=eq_dl,
                lhs_struct={"expr_type": "Class", "name": name},
                rhs_struct=eq_struct,
            )
            has_equiv = True

        if has_equiv:
            stats["defined_classes"] += 1
        else:
            stats["primitive_classes"] += 1

        for sup in cls.is_a:
            if sup is Thing:
                continue
            _add_subclass_decomposed(name, sup)

    for disj in onto.disjoint_classes():
        entities = list(disj.entities)
        if len(entities) < 2:
            continue
        names = sorted(_to_display_name(_clean_name(e.name)) for e in entities)
        for i in range(len(names)):
            for j in range(i + 1, len(names)):
                a, b = names[i], names[j]   # already lexicographic
                key = (a, b)
                if key in seen_disjoint_classes:
                    continue
                seen_disjoint_classes.add(key)
                add_axiom(
                    term_name=a,
                    term_type="Class",
                    axiom_type="DisjointClasses",
                    dl_str=f"DisjointClasses({a}, {b})",
                    subject=a,
                    object_=b,
                    lhs_struct={"expr_type": "Class", "name": a},
                    rhs_struct={"expr_type": "Class", "name": b},
                )


    for s, _, _ in rdf_graph.triples((None, OWL.disjointUnionOf, None)):
        union_node = list(rdf_graph.objects(s, OWL.disjointUnionOf))
        if not union_node:
            continue
        members = []
        cur = union_node[0]
        while cur and cur != RDF.nil:
            first = list(rdf_graph.objects(cur, RDF.first))
            if first:
                members.append(_to_display_name(_clean_name(str(first[0]))))
            rest = list(rdf_graph.objects(cur, RDF.rest))
            cur = rest[0] if rest else None
        cls_name = _to_display_name(_clean_name(str(s)))
        if not members:
            continue
        union_dl = " ⊔ ".join(members)
        add_axiom(
            term_name=cls_name,
            term_type="Class",
            axiom_type="DisjointUnion",
            dl_str=f"{cls_name} ≡ {union_dl}  (pairwise disjoint)",
            subject=cls_name,
            object_=union_dl,
            lhs_struct={"expr_type": "Class", "name": cls_name},
            rhs_struct={
                "expr_type": "DisjointUnion",
                "operands": [{"expr_type": "Class", "name": m} for m in members],
            },
        )


    for prop in onto.object_properties():
        name = _to_display_name(_clean_name(prop.name))
        stats["total_object_properties"] += 1

        add_axiom(
            term_name=name,
            term_type="ObjectProperty",
            axiom_type="Declaration",
            dl_str=f"Declaration: {name} (ObjectProperty)",
            subject=name,
            object_="ObjectProperty",
            lhs_struct={"expr_type": "ObjectProperty", "name": name},
            rhs_struct={"expr_type": "Declaration",
                        "term_type": "ObjectProperty"},
        )

        # SubPropertyOf
        for parent in prop.is_a:
            if not isinstance(parent, ObjectPropertyClass):
                continue
            if parent is prop:
                continue
            # Skip the synthetic top-level "ObjectProperty" pseudo-parent
            pname = _to_display_name(_clean_name(parent.name))
            if pname in ("ObjectProperty", "TopObjectProperty",
                         "topObjectProperty"):
                continue
            key = (name, "SubPropertyOf", pname)
            if key in seen_subprop:
                continue
            seen_subprop.add(key)
            add_axiom(
                term_name=name,
                term_type="ObjectProperty",
                axiom_type="SubPropertyOf",
                dl_str=f"{name} ⊑ {pname}",
                subject=name,
                object_=pname,
                lhs_struct={"expr_type": "ObjectProperty", "name": name},
                rhs_struct={"expr_type": "ObjectProperty", "name": pname},
            )

        # InverseOf
        if prop.inverse_property:
            inv = _to_display_name(_clean_name(prop.inverse_property.name))
            add_axiom(
                term_name=name,
                term_type="ObjectProperty",
                axiom_type="InverseOf",
                dl_str=f"{name} ≡ {inv}⁻",
                subject=name,
                object_=inv,
                lhs_struct={"expr_type": "ObjectProperty", "name": name},
                rhs_struct={"expr_type": "InverseObjectProperty", "name": inv},
            )

        # EquivalentProperties
        for eq_prop in getattr(prop, 'equivalent_to', []):
            eq_name = _to_display_name(_clean_name(getattr(eq_prop, 'name', str(eq_prop))))
            add_axiom(
                term_name=name,
                term_type="ObjectProperty",
                axiom_type="EquivalentProperties",
                dl_str=f"{name} ≡ {eq_name}",
                subject=name,
                object_=eq_name,
                lhs_struct={"expr_type": "ObjectProperty", "name": name},
                rhs_struct={"expr_type": "ObjectProperty", "name": eq_name},
            )

        # PropertyChain
        for chain in getattr(prop, 'property_chain', []):
            chain_parts = [_get_prop_name(p) for p in chain]
            chain_str = " ∘ ".join(chain_parts)
            add_axiom(
                term_name=name,
                term_type="ObjectProperty",
                axiom_type="PropertyChain",
                dl_str=f"{chain_str} ⊑ {name}",
                subject=name,
                object_=chain_str,
                lhs_struct={"expr_type": "ObjectProperty", "name": name},
                rhs_struct={
                    "expr_type": "ObjectPropertyChain",
                    "chain": [{"expr_type": "ObjectProperty", "name": p} for p in chain_parts],
                },
            )

        # Domain
        for d in prop.domain:
            dl = to_dl(d)
            add_axiom(
                term_name=name,
                term_type="ObjectProperty",
                axiom_type="Domain",
                dl_str=f"∃{name}.⊤ ⊑ {dl}",
                subject=name,
                object_=dl,
                lhs_struct={"expr_type": "ObjectProperty", "name": name},
                rhs_struct=expr_to_struct(d),
            )

        # Range
        for r in prop.range:
            dl = to_dl(r)
            add_axiom(
                term_name=name,
                term_type="ObjectProperty",
                axiom_type="Range",
                dl_str=f"⊤ ⊑ ∀{name}.{dl}",
                subject=name,
                object_=dl,
                lhs_struct={"expr_type": "ObjectProperty", "name": name},
                rhs_struct=expr_to_struct(r),
            )

        # Characteristics
        prop_iri = URIRef(prop.iri)
        for char_name in get_characteristics(prop_iri, rdf_graph):
            add_axiom(
                term_name=name,
                term_type="ObjectProperty",
                axiom_type="Characteristics",
                dl_str=f"{char_name}({name})",
                subject=name,
                object_=char_name,
                lhs_struct={"expr_type": "ObjectProperty", "name": name},
                rhs_struct={"expr_type": "PropertyCharacteristic", "name": char_name},
            )


    for prop in onto.data_properties():
        name = _to_display_name(_clean_name(prop.name))
        stats["total_datatype_properties"] += 1

        # Declaration (existence-only axiom)
        add_axiom(
            term_name=name,
            term_type="DatatypeProperty",
            axiom_type="Declaration",
            dl_str=f"Declaration: {name} (DatatypeProperty)",
            subject=name,
            object_="DatatypeProperty",
            lhs_struct={"expr_type": "DatatypeProperty", "name": name},
            rhs_struct={"expr_type": "Declaration",
                        "term_type": "DatatypeProperty"},
        )

        for parent in prop.is_a:
            if not isinstance(parent, DataPropertyClass):
                continue
            if parent is prop:
                continue
            pname = _to_display_name(_clean_name(parent.name))
            if pname in ("DatatypeProperty", "DataProperty",
                         "TopDataProperty", "topDataProperty"):
                continue
            key = (name, "SubPropertyOf", pname)
            if key in seen_subprop:
                continue
            seen_subprop.add(key)
            add_axiom(
                term_name=name,
                term_type="DatatypeProperty",
                axiom_type="SubPropertyOf",
                dl_str=f"{name} ⊑ {pname}",
                subject=name,
                object_=pname,
                lhs_struct={"expr_type": "DatatypeProperty", "name": name},
                rhs_struct={"expr_type": "DatatypeProperty", "name": pname},
            )

        for d in prop.domain:
            dl = to_dl(d)
            add_axiom(
                term_name=name,
                term_type="DatatypeProperty",
                axiom_type="Domain",
                dl_str=f"∃{name}.⊤ ⊑ {dl}",
                subject=name,
                object_=dl,
                lhs_struct={"expr_type": "DatatypeProperty", "name": name},
                rhs_struct=expr_to_struct(d),
            )

        for r in prop.range:
            dl = to_dl(r)
            add_axiom(
                term_name=name,
                term_type="DatatypeProperty",
                axiom_type="Range",
                dl_str=f"{name} maps-to {dl}",
                subject=name,
                object_=dl,
                lhs_struct={"expr_type": "DatatypeProperty", "name": name},
                rhs_struct=expr_to_struct(r),
            )

        prop_iri = URIRef(prop.iri)
        for char_name in get_characteristics(prop_iri, rdf_graph):
            add_axiom(
                term_name=name,
                term_type="DatatypeProperty",
                axiom_type="Characteristics",
                dl_str=f"{char_name}({name})",
                subject=name,
                object_=char_name,
                lhs_struct={"expr_type": "DatatypeProperty", "name": name},
                rhs_struct={"expr_type": "PropertyCharacteristic", "name": char_name},
            )

    # DisjointProperties
    for disj in onto.disjoint_properties():
        entities = list(disj.entities)
        if len(entities) < 2:
            continue
        # Determine if this is object or datatype side (mixed = fall back to ObjectProperty)
        is_data = any(isinstance(e, DataPropertyClass) for e in entities)
        prop_type = "DatatypeProperty" if is_data else "ObjectProperty"
        names = sorted(_to_display_name(_clean_name(e.name)) for e in entities)
        for i in range(len(names)):
            for j in range(i + 1, len(names)):
                a, b = names[i], names[j]
                key = (prop_type, a, b)
                if key in seen_disjoint_props:
                    continue
                seen_disjoint_props.add(key)
                add_axiom(
                    term_name=a,
                    term_type=prop_type,
                    axiom_type="DisjointProperties",
                    dl_str=f"DisjointProperties({a}, {b})",
                    subject=a,
                    object_=b,
                    lhs_struct={"expr_type": prop_type, "name": a},
                    rhs_struct={"expr_type": prop_type, "name": b},
                )

    if temp_to_cleanup and os.path.exists(temp_to_cleanup):
        try:
            os.unlink(temp_to_cleanup)
        except OSError:
            pass

    return terms, stats, name_to_label


def _sanitize_for_excel(value):

    if value is None:
        return ""
    s = str(value)
    out_chars = []
    for ch in s:
        cp = ord(ch)
        if cp == 0xFEFF:                     
            continue
        if cp < 0x20 and cp not in (0x09, 0x0A, 0x0D):
            continue
        if cp == 0x7F:                        
            continue
        out_chars.append(ch)
    out = "".join(out_chars)
    if out and out[0] in ("=", "+", "-", "@"):
        out = "'" + out
    return out


def save_excel(terms, excel_path, name_to_label=None):
    from openpyxl import Workbook
    name_to_label = name_to_label or {}
    wb = Workbook()
    ws = wb.active
    ws.title = "Atomic TBox Axioms"

    headers = ["Term", "TermLabel", "TermType", "AxiomType",
               "Subject", "SubjectLabel", "Object", "DL"]
    ws.append(headers)

    header_fill = PatternFill("solid", fgColor="4472C4")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    def _label_for(name):
        return name_to_label.get(name, name)

    for term_name, entry in sorted(terms.items()):
        for ax in entry["axioms"]:
            subj = ax.get("subject", "")
            ws.append([
                _sanitize_for_excel(term_name),
                _sanitize_for_excel(_label_for(term_name)),
                _sanitize_for_excel(entry["term_type"] or ""),
                _sanitize_for_excel(ax["axiom_type"]),
                _sanitize_for_excel(subj),
                _sanitize_for_excel(_label_for(subj)),
                _sanitize_for_excel(ax.get("object", "")),
                _sanitize_for_excel(ax.get("dl", "")),
            ])

    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 80)

    ws.freeze_panes = "A2"
    wb.save(excel_path)
    print(f"Excel saved to: {excel_path}")



def save_json(terms, stats, json_path, name_to_label=None):
    name_to_label = name_to_label or {}

    def _label_for(name):
        return name_to_label.get(name, name)

    type_order = {"Class": 0, "ObjectProperty": 1, "DatatypeProperty": 2}

    flat = []
    items = sorted(
        terms.items(),
        key=lambda kv: (type_order.get(kv[1]["term_type"], 99), kv[0]),
    )
    for term_name, entry in items:
        for ax in entry["axioms"]:
            subj = ax.get("subject", "")
            flat.append({
                "term": term_name,
                "term_label": _label_for(term_name),
                "term_type": entry["term_type"],
                "axiom_type": ax["axiom_type"],
                "subject": subj,
                "subject_label": _label_for(subj),
                "object": ax.get("object", ""),
                "dl": ax.get("dl", ""),
                "lhs_struct": ax.get("lhs_struct"),
                "rhs_struct": ax.get("rhs_struct"),
            })


    flat.sort(key=lambda a: (
        type_order.get(a["term_type"], 99),
        a["term"],
        a["axiom_type"],
    ))

    for i, ax in enumerate(flat, start=1):
        ax_with_id = {"id": f"AX{i}"}
        ax_with_id.update(ax)
        flat[i - 1] = ax_with_id

    output = {
        "stats": dict(stats),
        "axioms": flat,
    }
    if name_to_label:
        output["name_to_label"] = dict(name_to_label)

    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(output, f, ensure_ascii=False, indent=2)
    print(f"JSON saved to: {json_path}")



def save_txt(terms, stats, txt_path):
    lines = []
    add = lambda text="": lines.append(text)
    sep = "=" * 100

    add(sep)
    add("ATOMIC TBOX AXIOMS")
    add(sep)
    add()

    groups = {"Class": [], "ObjectProperty": [], "DatatypeProperty": []}
    for name, entry in sorted(terms.items()):
        tt = entry["term_type"]
        if tt in groups:
            groups[tt].append((name, entry))

    for sect, items in groups.items():
        if not items:
            continue
        add(f"── {sect} ──")
        add()
        for term_name, entry in items:
            add(f"  Term: {term_name}  [{entry['term_type']}]")
            for ax in entry["axioms"]:
                add(f"    {ax['axiom_type']:<25}  {ax['dl']}")
                add(f"      Subject: {ax.get('subject', '')}")
                add(f"      Object : {ax.get('object', '')}")
            add()
        add()

    add(sep)
    add("STATISTICS")
    add(sep)
    add()
    add(f"  Classes:               {stats.get('total_classes', 0)}  "
        f"(defined: {stats.get('defined_classes', 0)}, "
        f"primitive: {stats.get('primitive_classes', 0)})")
    add(f"  ObjectProperties:      {stats.get('total_object_properties', 0)}")
    add(f"  DatatypeProperties:    {stats.get('total_datatype_properties', 0)}")
    add()
    add("  Axiom Breakdown:")
    skip_keys = {"total_classes", "defined_classes", "primitive_classes",
                 "total_object_properties", "total_datatype_properties",
                 "total_axioms"}
    for k in sorted(stats.keys()):
        if k in skip_keys:
            continue
        add(f"    {k:<25} {stats[k]}")
    add(f"    {'-' * 30}")
    add(f"    {'Total TBox Axioms':<25} {stats.get('total_axioms', 0)}")
    add()
    add(sep)

    text = "\n".join(lines)
    with open(txt_path, "w", encoding="utf-8") as f:
        f.write(text)
    print(f"TXT saved to: {txt_path}")
    return text


if __name__ == "__main__":
    if len(sys.argv) != 2:
        print("Usage: python extract_atomic_tbox.py <ontology_path>")
        sys.exit(1)

    ontology_path = sys.argv[1]
    base = ontology_path.rsplit('.', 1)[0]
    output_txt = base + "_atomic_tbox.txt"
    output_json = base + "_atomic_tbox.json"
    output_excel = base + "_atomic_tbox.xlsx"

    print(f"Loading ontology: {ontology_path}")
    try:
        terms, stats, name_to_label = extract_tbox_axioms(ontology_path)
        save_txt(terms, stats, output_txt)
        save_json(terms, stats, output_json, name_to_label=name_to_label)
        save_excel(terms, output_excel, name_to_label=name_to_label)
        print("Done.")
    except Exception as e:
        import traceback
        traceback.print_exc()
        sys.exit(1)
